"""
Copyright 2018 IBM Corporation
Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

"""
 We need to pass through content at least twice to connect label with node_names, we use third pass to generate XML
  The first pass: (processing separately each file)
     - generating bag of blocks. Block is a segment of rows separated by empty line,
           it is used to generate intent, entity, or node or combination of two definitions (intent-node, entity-node)
           it can generate a sequence of nodes (without intent or entity) if header with X_PLACEHOLDER,
     - creating list of labels (in DialogData) (but not assigning node IDs yet)

  The second pass: (processing all files together)
       - populating DialogData by Intent definitions (IntentData), Node definitions (NodeData) and entity definitions (EntityData).
       - populating node_names and mapping label-nodeName  
  The third pass:  Generating XML, CSV  - XML per original document - this is done in XMLHandler 
       - generate jump-to based on label map

  Rational: what are the dependences
    intent_name is derived from header or firt example if the construct defines an intent (starts with #header or without it
       - if it is auto generated - it should be made unique across all the workspaces
       - if it is assigned - it is only checked if it is unique
   condition - makes sense only if output is provided. Conditions are not unique. They are base for generation of unique node_names
   node_name - is derived from condition (unless we are using meta command to assign it)
       - they should be unique across all node names, if already exists - we ad _xxx where xxx is a unique number
"""
import os, re
import unicodedata, unidecode
from openpyxl import load_workbook
from wawCommons import printf, eprintf, toIntentName
from zipfile import BadZipfile
from xml.sax.saxutils import escape
import DialogData as Dialog
from DialogData import DialogData
from itertools import izip_longest
from NodeData import NodeData
import more_itertools


class XLSXHandler(object):
    """ Converts Excel spreadsheet from multiple files to an internal data representation in DialogData.
    """

    def __init__(self, config):
        self._blocks = []  # internal representation of XLS, list of blocks, (block are the lines separated by empty line)
        self._dialogData = DialogData(config)  # internal representation of the workspace
        self._nodeData = NodeData()
        self._config = config  # we need config to get NAME_POLICY, verbosity,..
        self._VERBOSE = hasattr(config, 'common_verbose')
        self._NAME_POLICY = 'soft'  # TBD: enable to set the NamePolicy from config file
        self.menu_reacts = []   #item from dialog data
        self.menu_blocks = []  # menu blocks
        self._numerized_outputs = []  # create numerized mp3
        self._num_of_options=[]     #number of options in menu
        self._cond_to_menu=[]
        self._dictionary={}


    def getBlocks(self):
        return self._blocks

    def addBlock(self, dataBlock):
        self._blocks.append(dataBlock)

    def getDialogData(self):
        return self._dialogData

    def parseXLSXIntoDataBlocks(self, filename):
        """ Reads Excel spreadsheet (in T2C format). Splits it to blocks and
            stores the data as tuples (domain, prefix, intent, rawBlock) in _dataBlocks,
            THIS IS THE FIRST PASS THROUGH INPUT  (a single file of the INPUT)
        """
        printf('Processing xlsx file: %s\n', filename)
        if not os.path.exists(filename):
            eprintf('Error: File does not exist: %s\n', filename)
            return {}

        # Derive domain name from file name (use the same naming policy as for intents)
        try:
            domainName = toIntentName(self._NAME_POLICY, None, os.path.splitext(os.path.split(filename)[1])[0])
            try:
                domainName = unicode(domainName, 'utf-8')  # Python 2
            except NameError:
                domainName = str(domainName)  # Python 3
            workbook = load_workbook(filename=filename, read_only=True)
        except (IOError, BadZipfile):
            eprintf('Error: File does not seem to be a valid Excel spreadsheet: %s\n', filename)
            return {}

        # Process all the tabs of the file
        for sheet in workbook.worksheets:
            # get prefix is a sheet title
            printf(' Sheet: %s\n', sheet.title)
            try:
                prefix = (sheet.title.encode('utf-8'))  # Python 2
            except NameError:
                prefix = str(sheet.title)  # Python 3

            currentBlock = []  # Each cheet starts a new block
            # Separate all data blocks in the sheet, if the currentBlock starts with header, the header is considered to be part of currentBlock
            for row in sheet.iter_rows(max_col=4):
                validRow = False
                # Check if the row is valid. Row is valid if it contains at least one column not empty and different from comment
                for columnIndex in range(0, 4):
                    if row[columnIndex] and row[columnIndex].value and not (row[columnIndex].value.startswith('//')):
                        validRow = True

                # Three slashes in the first cell cause whole rest of the line to be treated as comment
                if row[0].value and row[0].value.startswith('///'):
                    validRow = False

                if not validRow:
                    # If behind the block, we save the currentBlock (if any was populated)
                    if currentBlock:
                        self.__createBlock(domainName, prefix, currentBlock)
                    currentBlock = []
                else:
                    # if valid row - we add the raw to block
                    currentBlock.append(
                        (escape(row[0].value.strip()) if row[0].value and not row[0].value.startswith('//') else None,
                         escape(row[1].value.strip()) if row[1].value and not row[1].value.startswith('//') else None,
                         escape(row[2].value.strip()) if row[2].value and not row[2].value.startswith('//') else None,
                         escape(row[3].value.strip()) if row[3].value and not row[3].value.startswith('//') else None))
            if currentBlock:
                self.__createBlock(domainName, prefix, currentBlock)  # store the last block of the sheet

    def __createBlock(self, domain, prefix, block):
        """ Add the block to the block list """
        if not block or not block[0][0]:
            printf('WARNING: First cell of the data block does not contain any data. (domain=%s, prefix=%s)\n', domain,
                   prefix)
            if block[0][1]:
                self.menu_reacts.append(block[0][1])
            elif block[0][2]:
               pass
            else:
                return
        self._blocks.append((domain, prefix, block))


    def __is_condition_block(self, block):
        """ Returns true if first cell contains X_PLACEHOLDER
            or more then 1 condition indicator (one is just a header)
        """
        no_special = len(re.sub('[^#$@&|]', '', block[0][0]))
        if (Dialog.X_PLACEHOLDER in block[0][0] and not block[0][1]):  # header containing X_PLACEHOLDER
            return True
        if ((no_special > 0) and block[0][1]):  # Simple intent-output pair,
            return True  # simple intent-output pair,
        return False

    def __is_header(self, block):
        """ :returns true if the block has a header i.e. first line has col1 but not col2"""
        return block[0][0] and not (block[0][1])

    def __separate_label_from_block(self, block):
        """ If block has a label (starts with : e.g.  :xxxx) it removes it from the block
            :returns label or None (if no label was found )
        """
        label = None
        test_val = []
        firstCell = block[0][0]  # firstCell is a header, condition
        if firstCell.startswith("!"):
            if "!Def" in firstCell:
                # print type(firstCell)
                self._dialogData._arduino_definitions.append(firstCell)
            if "!Act" in firstCell:
                self._dialogData._arduino_definitions.append(firstCell)
        if firstCell.startswith(u':') and len(block[0][0]) > 1:
            label = firstCell[1:]
            if self._dialogData.isLabel(label):
                printf(
                    'WARNING: Found a label that has already been assigned to an intent and will be overwritten. Label: %s\n',
                    label)
            del block[0]  # delete line with label
            if not block or not firstCell:
                printf(
                    'ERROR: First cell of the goto block does not contain any data. \n')
                exit()
        return label

    def updateReferences(self):
        """
              Replaces labels by node_names  (second pass)
        """
        self._dialogData.updateReferencesNodes()

    def convertBlocksToDialogData(self):
        """ Reads all blocks to _dataBlocks and handles each depending on the block type.
            This is typically called only once after all T2C data are processed and after they are put to _dataBlocks
             by parseXLSXIntoDataBlocks
        """
        for domain, prefix, block in self._blocks:  # For each block
            # Validity check of parameters
            if not block or not isinstance(block[0], tuple) or not block[0][0]:
                continue
            if block[0][1] and not block[0][0]:
                continue
            # separate label, strip it from block
            label = self.__separate_label_from_block(block)
            firstCell = block[0][0]
            # Block has header if it starts only with column 1 cell (no other cells)!
            #   It can not be implicit intent definition as it need an output, otherwise noone coud referr to it
            blockHasHeader = self.__is_header(block)
            conditionHasX = Dialog.X_PLACEHOLDER in firstCell
            if conditionHasX and not blockHasHeader:
                printf('WARNING: Value next to header. (domain=%s, prefix=%s, row=%s)\n', domain, prefix, block[0])
                exit()

            if self.__is_condition_block(block):  # Condition block does not define an intent nor entity,
                conditionHasX = Dialog.X_PLACEHOLDER in block[0][0]
                if not conditionHasX:  # simple pair condition-output, condition is one, output can have more outputs
                    # left column is a condition
                    self.__handle_condition_block(block, domain, label)
                else:
                    # x_condition block has header with templateor, filler to a X in teplate and bunch of outputs for each filler
                    self.__handle_x_condition_block(block, domain, label)
            else:
                if firstCell.startswith(u'@'):
                    # The entity block includes definition of intent (it contains input value/synonym list)
                    #  it must start with header,
                    #  it can have outputs (header is still required, otherwise it wold be mistakable for entity)
                    self.__handle_entity_block(block, domain, label)
                else:
                    # The intent block includes definition of intent (it contains input example list)
                    #  it can start with header,
                    #  it can have outputs (then header is not required)
                    #  if it has outputs, it does not need a header
                    self.__handle_intent_block(block, domain, label)

    def __handle_x_condition_block(self, block, domain, label):
        """ Handles x conditional blocks.
            x condition block has header with template, fillers (in column 1 below header) and bunch of outputs for each filler
            Node conditions are obtained by replacing <x>  by filler from the first column
       """
        if block[0][1]:
            eprintf('ERROR: Found occupied cell after the header of conditional block: %s\n', block[0])
            exit()
        if not block[1][0]:
            eprintf('ERROR: Header with X is not followed by any value %s\n', block[0])
            exit()
        if not block[1][1]:
            eprintf('ERROR: Header with X should be followed by at least one output %s\n', block[0])
            exit()

        # there are as many nodes as outputs
        block_header = block[0][0]
        for row in block[1:]:
            node_condition = re.sub(Dialog.X_PLACEHOLDER, row[0], block_header)
            node_name = self._dialogData.createUniqueNodeName(node_condition)  # make it unique
            nodeData = self._dialogData.createNode(node_name, domain)  # create space for new node, remembers node_name
            nodeData.setCondition(node_condition)
            if row[1]:
                nodeData.addRawOutput(row[1:], self._dialogData.getLabelsMap())
            else:
                eprintf('Warning: Format error, no output defined for the condition :%s', row)


    def __handle_condition_block(self, block, domain, label):
        """ Handles simple pair condition-output,
              condition is one, output can have more outputs
        """
        if not (block[0][1]):
            eprintf('ERROR: Format error. condition does not have a output : %s\n', block[0])
            exit()
        if len(block) > 1 and block[1][0]:
            eprintf('ERROR: Format error. condititional block without header shold have just one condition: %s\n',
                    block[0])
            exit()
        node_name = self._dialogData.createUniqueNodeName(
            block[0][0])  # derive node name from explicit condition and make it unique
        node_condition = block[0][0]
        nodeData = self._dialogData.createNode(node_name, domain)  # create space for new node, remembers node_name
        # nodeData.setName(node_name)
        nodeData.setCondition(node_condition)

        for row in block:
            if row[1]:
                nodeData.addRawOutput(row[1:], self._dialogData.getAllEntities())
            else:
                eprintf(
                    'ERROR: Format error. empty output in a conditional block does not make sense. %s\n', block[0])
                exit()
        if label:  # add lable - if any
            self._dialogData.addLabel(label, node_name)

    def __handle_entity_block(self, block, domain, label):
        """ Proess entity definition with possible output
            - block can start with header,
            - block can have output(s), then header is not required
        """
        startsWithHeader = block[0][0].startswith(u'@')  # Is it header? If so, it starts with @
        if not startsWithHeader and not block[0][1]:
            eprintf('ERROR: Internal error. __handle_entity_block handling ConditionBlock : %s\n', block[0])
            exit()
        if startsWithHeader:
            entity_name = block[0][0][1:]  # header is name (including hash)
        else:
            # if no name -derive intent name automatically form the first sentence, make it unique among intents
            entity_name = self._dialogData.createUniqueEntityName(block[0][0])

        entityData = self._dialogData.createEntity(entity_name)  # create space for new entity
        first_output = block[1][1] if startsWithHeader else block[0][
            1]  # if we have a header, the first output is in second row
        if first_output:  # if first output then any output -> we generate a node, assign label ..
            node_name = self._dialogData.createUniqueNodeName(
                entity_name)  # derive node name from explicit intent name, make it unique
            nodeData = self._dialogData.createNode(node_name, domain)  # create space for new node, remembers node_name
            # nodeData.setName(node_name)
            node_condition = entity_name
            nodeData.setCondition(node_condition)
            if label:
                self._dialogData.addLabel(label, node_name)
        else:  # only entity definition
            if label:  # this block does not generate node, label does no make sense
                eprintf('ERROR: Format error. Label is next to the block which is not generating node : %s\n', block[0])
                exit()
        for row in block:
            if row[0] and not row[0].startswith(u'@'):  # we skip header if any
                entityData.addValue(row[0].rstrip().rstrip(';'))  # Collect entity values and synonyms
            if row[1]:
                if not first_output:
                    eprintf('ERROR: Format error. Adjacent outputs are not in a sigle block : %s\n', block[0])
                    exit()
                else:
                    nodeData.addRawOutput(row[1:], self._dialogData.getAllEntities())

    def __handle_intent_block(self, block, domain, label):
        """ Process intent definition optionally also with output
            - block has optional header defining the name of the intent (header is needed if no output is defined
              (reason - difficult to refer to the intent),
            - block has optional output(s) (then header is not required but also optional)
        """

        startsWithHeader = block[0][0].startswith(u'#')  # is header?
        if block[0][0].startswith('!') and not block[0][1]:
            pass
        # elif block[0][0] == '!!Menu' and not block[0][1]:
        #     pass
        elif not startsWithHeader and not block[0][1]:
            eprintf('ERROR: Internal error. __handle_intent_block handling ConditionBlock : %s\n', block[0])
            exit()
        if startsWithHeader:
            intent_name = block[0][0][1:]  # header is name (including hash)
        else:
            # if no name -derive intent name automatically form the first sentence, make it unique among intents
            intent_name = self._dialogData.createUniqueIntentName(block[0][0])

        intentData = self._dialogData.createIntent(intent_name)  # create space for new intent

        first_output = block[1][1] if startsWithHeader else block[0][
            1]  # if we have a header, the first output is in second row
        if first_output:  # if first output then any output -> we generate a node, assign label ..
            node_name = self._dialogData.createUniqueNodeName(
                intent_name)  # derive node name from explicit intent name, make it unique
            nodeData = self._dialogData.createNode(node_name, domain)  # create space for new node, remembers node_name
            # nodeData.setName(node_name)  #- not needed- set by createNode
            node_condition = '#' + intent_name
            nodeData.setCondition(node_condition)

            if label:
                self._dialogData.addLabel(label, node_name)
        else:  # only intent definition
            if label:  # this block does not generate node, label does no make sense
                eprintf('ERROR: Format error. Label is next to the block which is not generating node : %s\n', block[0])
                exit()

        for row in block:
            if row[0] and not row[0].startswith(u'#'):  # we skip header if any
                intentData.addExample(row[0])  # Collect intent definition
            if row[1]:
                if block[0][0].startswith('!Menu') and not first_output:
                    print(row[1])
                    continue
                elif not first_output:
                    eprintf('ERROR: Format error. Adjacent outputs are not in a sigle block : %s\n', block[0])
                    exit()
                else:
                    nodeData.addRawOutput(row[1:], self._dialogData.getAllEntities())
                    print(row[1])
    def concatenate_menus_to_list(self):
        # find menu in blocks and make list of them
        numerized_outputs=self._numerized_outputs
        blocks = self._blocks
        for block in blocks:
            if block[2][0][1] and len(block[2]) == 1:
                if not block[2][0][0]:
                    self._dialogData._menu.append(self.menu_blocks[0])
                    self.menu_blocks.pop(0)
                    self._dialogData._menu.append(len(block[2]))
            if block[2][0][0] and not block[2][0][1]:
                if '!Menu' in block[2][0][0]:
                    self._dialogData._menu.append(block[2][0][0])
            elif len(block[2]) > 1:
                if not block [2][0][0]:
                    self._dialogData._menu.append(self.menu_blocks[0])
                    self.menu_blocks.pop(0)
                    self._dialogData._menu.append(len(block[2]))
        # print self._dialogData.get_menu()
    def create_numerized_outputs(self):
        menu_reacts=self.menu_reacts
        # print menu_reacts
        outputs = self._numerized_outputs
        blocks = self._blocks
        menu = self.menu_blocks
        num_of_options=self._num_of_options
        complete_list_of_all_outputs=self._dialogData._all_menu_outputs
        # CREATING A LIST OF A SECOND COLUMN IN EXCEL SHEET
        for block in blocks:
            if block[2][0][1] and len(block[2]) == 1:
                outputs.append(block[2][0][1])
                complete_list_of_all_outputs.append(block[2][0][1])
                if block[2][0][1] and not block[2][0][0]:
                    num_of_options.append(len(block[2]))

            elif len(block[2]) > 1:
                for i in range(len(block[2])):
                    outputs.append(block[2][i][1])
                    complete_list_of_all_outputs.append(block[2][i][1])
                if block[2][0][1] and not block[2][0][0]:
                    num_of_options.append(len(block[2]))

        # FINDING THE ITEMS OF THE MENU
        # print outputs
        index_of_match = [x for x,item in enumerate(outputs) if item in menu_reacts]
        for n, item in enumerate(outputs):
            num = str(n + 1).zfill(3)
            outputs[n] = num
            # ADDING THE ENUMERATED OUTPUTS TO MENU
            if n in index_of_match:
                menu.append(num)
                # menu.insert(-1,num_of_options[0])
                num_of_options.pop(0)

        # print outputs
        self.concatenate_menus_to_list()
        return outputs

    def create_reactive(self):
        numerized_outputs = self._numerized_outputs
        conditions = self._cond_to_menu
        res=[]
        index=0
        num=0
        for n,block in enumerate(self._blocks):
            if block[2][0][0] and block[2][0][0].startswith("!"):
                if "!Menu" in block[2][0][0]:
                    break
                res.append("CUT")
            if len(block[2])==1:
                if block[2][0][1] and not "!Menu" in block[2][0][0]:
                    res.append(conditions[num])
                    res.append(numerized_outputs[index])
                    res.append(len(block[2]))
                    index+=1
                    num+=1
            else:
                for i in range (len(block[2])):
                    if block[2][i][0]:
                        res.append(conditions[num])
                        res.append(numerized_outputs[index])
                        res.append(len(block[2]))
                        index+=len(block[2])
                        num+=1
        reactives = list(more_itertools.split_at(res, lambda s: s == 'CUT'))
        reactives=[x for x in reactives if len(x)>1]
        # print len(reactives)
        for react in reactives:
            self._dialogData._reactive_outputs.append(react)
        print self._dialogData.get_reactive_outputs()

    def menu_handling(self, block):
        numerized_outputs = self.create_numerized_outputs()
        menu = self._dialogData.get_menu()
        # print menu
        # divide all menus to individual lists
        menus = list(more_itertools.split_after(menu, lambda s: s == u'!!Menu'))
        # print menus
        # CREATING A WORKSPACE FOR MENU
        for menu in menus:
            # print menu
            menu_workspace=str(menu)
            # print menu_workspace
            param1 = 0
            for ch in ['[',']',"'"]:
                if ch in menu_workspace:
                    menu_workspace=menu_workspace.replace(ch,"")
            menu_workspace=menu_workspace.replace('u!', '!')
            # print menu_workspace
            name_of_menu = menu_workspace[6:menu_workspace.find(';')]
            index_round_flat=6
            if 'auto'.upper() in name_of_menu:
                param1+=0x01
            if 'intro'.upper() in name_of_menu:
                param1+=0x02
            if 'flat' in name_of_menu: #CHECKING FLAT OR ROUND PARAMETR
                index_round_flat=(name_of_menu.index('flat'))
            elif 'round' in name_of_menu:
                index_round_flat =name_of_menu.index('round')
                param1+=0x10
            starting_index=6+len(name_of_menu)+1
            name_of_menu=name_of_menu[:index_round_flat].strip()  # DEFINING A NAME OF THE MENU
            menu_workspace=menu_workspace[starting_index::]
            order=menu_workspace[6:menu_workspace.index(';')]
            # CHECKING AN ORDER OF THE MENU
            if "last" in order:
                param1+=0x04
            if "first" and "last" not in order:
                param1+=0x08
            starting_index=6+len(order)+1
            menu_workspace=menu_workspace[starting_index::]
            timeout=menu_workspace[8:menu_workspace.index(',')]       # DEFINING TIMEOUT OF THE MENU
            starting_index=8+len(timeout)+1
            ending_index=menu_workspace.index("!!Menu")
            menu_workspace=menu_workspace[starting_index:ending_index-2]
            # THE FINAL LOOK OF MENU FOR ARDUINO
            final_menu=name_of_menu+"[]"+'{4, '+str(param1)+', '+'10'+', '+timeout+', '+menu_workspace+'}'
            self._dialogData._all_menu.append(final_menu)
            # self.create_reactive()

    def definition_handler(self):
        blocks=self._blocks
        keys = []
        vals = []
        dictionary = self._dictionary
        definitions = self._dialogData.get_arduino_definitions()
        for item in definitions:
            lst = filter(None, re.split("[= ]", item))
            if '#' in lst[1]:
                keys.append(lst[1])
                vals.append(int(lst[2]))
        dictionary = dict(zip(keys, vals))
        for block in blocks:
            if block[2][0][0] in dictionary.keys():
                self._cond_to_menu.append(dictionary[block[2][0][0]])










